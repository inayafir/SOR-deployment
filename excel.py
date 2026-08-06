"""Excel import/export for the SOR database.

Import expects a spreadsheet with (at least) columns named like
``sor number``, ``title``, ``category`` and ``price``. Column headers are
matched case-insensitively and ignoring separators, so common variants
(``SOR Code``, ``Service Name``, ``Rate (Rs)``) are accepted. A bulk import
validates the whole file first and replaces the live dataset only when every
row is valid — nothing is half-imported.
"""

import io
import math
import re
from datetime import datetime

import openpyxl

import db
from categories import classify_category

ALPHANUMERIC_RE = re.compile(r"^[A-Za-z0-9]+$")

_CODE_CANDIDATES = {"sornumber", "sorno", "sor", "sorcode", "code", "slno"}
_TITLE_CANDIDATES = {
    "title", "name", "servicename", "service", "procedure",
    "servicetitle", "treatment",
}
_CATEGORY_CANDIDATES = {"category", "cat", "speciality", "specialty"}
_PRICE_CANDIDATES = {"price", "rate", "amount", "cost", "chssrate", "newrate", "rateinr"}


def _norm_header(value):
    return re.sub(r"[^a-z0-9]", "", (value or "").lower())


def _find_column(headers, candidates):
    for index, header in enumerate(headers):
        if _norm_header(header) in candidates:
            return index
    return None


def _parse_price(value):
    """Parse a price cell. Blank -> None; number -> float; annotated text
    follows the app's convention (first number found, else None for values
    like 'At actuals')."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        price = float(value)
        if not math.isfinite(price) or price < 0:
            raise ValueError("price must be a non-negative number")
        return price
    text = str(value).strip()
    if not text:
        return None
    return db._parse_rate(text)


def parse_import_rows(source):
    """Validate and parse an Excel import file.

    ``source`` may be raw ``bytes``, a seekable file-like object, or a
    filesystem path. Returns a list of dicts with keys ``sor_code``,
    ``name``, ``category`` and ``price``, ready for
    :func:`db.replace_sor_items`. Raises ``ValueError`` with a
    human-readable message when the file is malformed or any row is
    invalid, in which case the live database is untouched.
    """
    if isinstance(source, bytes):
        source = io.BytesIO(source)
    try:
        workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(
            "Could not open the file as an Excel workbook. "
            "Please upload a valid .xlsx file."
        ) from exc

    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ValueError("The Excel file is empty.")
        headers = list(header)

        code_col = _find_column(headers, _CODE_CANDIDATES)
        title_col = _find_column(headers, _TITLE_CANDIDATES)
        category_col = _find_column(headers, _CATEGORY_CANDIDATES)
        price_col = _find_column(headers, _PRICE_CANDIDATES)

        if title_col is None:
            raise ValueError(
                "Could not find a 'title' column. "
                "Expected columns: sor number, title, category, price."
            )

        errors = []
        items = []
        seen_codes = set()
        row_number = 1

        def cell(values, col):
            if col is None or col >= len(values):
                return None
            return values[col]

        for values in rows:
            row_number += 1
            values = list(values) if values is not None else []

            raw_code = cell(values, code_col)
            raw_title = cell(values, title_col)
            raw_category = cell(values, category_col)
            raw_price = cell(values, price_col)

            code = str(raw_code).strip() if raw_code is not None else ""
            title = str(raw_title).strip() if raw_title is not None else ""

            if not code and not title:
                continue

            if not title:
                errors.append(f"Row {row_number}: title is required.")
                continue

            if code:
                if not ALPHANUMERIC_RE.fullmatch(code):
                    errors.append(
                        f"Row {row_number}: SOR number '{code}' must be "
                        "alphanumeric (letters and digits only)."
                    )
                if code in seen_codes:
                    errors.append(
                        f"Row {row_number}: duplicate SOR number '{code}'."
                    )
                seen_codes.add(code)

            category = str(raw_category).strip() if raw_category is not None else ""
            if not category:
                category = classify_category(title, code)

            try:
                price = _parse_price(raw_price)
            except ValueError:
                errors.append(
                    f"Row {row_number}: price must be a non-negative number."
                )
                continue

            items.append({
                "sor_code": code or None,
                "name": title,
                "category": category,
                "price": price,
            })

        if errors:
            shown = errors[:10]
            summary = "; ".join(shown)
            more = (
                f" (and {len(errors) - len(shown)} more)"
                if len(errors) > len(shown)
                else ""
            )
            raise ValueError(
                f"The Excel file could not be imported: {summary}{more}"
            )
        if not items:
            raise ValueError("The Excel file has no SOR items to import.")
        return items
    finally:
        workbook.close()


def export_to_bytes():
    """Build an Excel workbook of the current database and return its bytes."""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "SOR"
    sheet.append(["SOR Number", "Title", "Category", "Price"])
    for item in db.list_all_sor_items():
        sheet.append([
            item["sor_code"],
            item["name"],
            item["category"],
            item["price"] if item["price"] is not None else None,
        ])
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def default_filename():
    return f"sor_database_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
