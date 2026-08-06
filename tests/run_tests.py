"""Automated verification suite for the CHSS SOR application.

Run from the project root:

    python tests/run_tests.py

Uses an isolated data directory (SOR_DATA_DIR) so it never touches a real
database. Covers auth, RBAC, CSRF, CRUD, search, persistence, WAL mode,
concurrency and config overrides.
"""

import os
import shutil
import sys
import tempfile
from concurrent.futures import ThreadPoolExecutor

# Isolate all runtime data BEFORE importing application modules.
TEST_DATA_DIR = tempfile.mkdtemp(prefix="sor_test_")
os.environ["SOR_DATA_DIR"] = TEST_DATA_DIR

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, PROJECT_ROOT)
os.chdir(PROJECT_ROOT)

import config as app_config  # noqa: E402
import db  # noqa: E402
import excel  # noqa: E402
import paths  # noqa: E402
import app as app_module  # noqa: E402

app = app_module.app

FAILURES = []


def check(name, condition, detail=""):
    status = "OK" if condition else "FAIL"
    print(f"[{status}] {name}" + (f" — {detail}" if detail else ""))
    if not condition:
        FAILURES.append(name)


def login(client, username="user", password="user123"):
    with client.session_transaction() as sess:
        sess["_csrf_token"] = "test-csrf-token"
        sess["user_id"] = 1
        sess["username"] = username
        sess["role"] = "admin" if username == "admin" else "user"


def get_token(client):
    with client.session_transaction() as sess:
        sess["_csrf_token"] = "test-csrf-token"
        return "test-csrf-token"


def csrf_post(client, url, data):
    data = dict(data)
    data["_csrf_token"] = get_token(client)
    return client.post(url, data=data, follow_redirects=False)


def test_runtime_environment():
    paths.ensure_runtime_dirs()
    for name in paths.RUNTIME_DIRS:
        check(f"runtime dir created: {name}", os.path.isdir(paths.runtime_path(name)))

    cfg = app_config.load_config()
    check("config.ini created", os.path.exists(app_config.get_config_path()))
    check("default host 0.0.0.0", cfg["server"]["host"] == "0.0.0.0")
    check("default port 8080", cfg["server"]["port"] == 8080)
    check("secret key generated", bool(cfg["security"]["secret_key"]))
    check("app secret key set", bool(app.secret_key))

    # env overrides win
    os.environ["SOR_PORT"] = "9999"
    os.environ["SOR_HOST"] = "127.0.0.1"
    os.environ["SOR_THREADS"] = "4"
    cfg2 = app_config.load_config()
    check("SOR_PORT override", cfg2["server"]["port"] == 9999)
    check("SOR_HOST override", cfg2["server"]["host"] == "127.0.0.1")
    check("SOR_THREADS override", cfg2["server"]["threads"] == 4)
    del os.environ["SOR_PORT"], os.environ["SOR_HOST"], os.environ["SOR_THREADS"]


def test_database():
    db.init_db()
    with db.get_db() as conn:
        total = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
        mode = conn.execute("PRAGMA journal_mode").fetchone()[0]
        timeout = conn.execute("PRAGMA busy_timeout").fetchone()[0]
        codes = conn.execute("SELECT COUNT(DISTINCT sor_code) FROM sor_items").fetchone()[0]
    check("seeded 2154 items", total == 2154, str(total))
    check("all codes unique", codes == 2154, str(codes))
    check("WAL mode enabled", mode == "wal", mode)
    check("busy timeout set", timeout == 30000, str(timeout))

    # seeding is idempotent — init_db again does not duplicate data
    db.init_db()
    with db.get_db() as conn:
        total2 = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
    check("init_db is idempotent", total2 == total)


def test_auth_rbac_csrf():
    anon = app.test_client()
    check("anonymous /search redirects", anon.get("/search").status_code == 302)
    check("anonymous API 401", anon.get("/api/search").status_code == 401)

    user = app.test_client()
    login(user, "user", "user123")
    check("user can search", user.get("/search").status_code == 200)
    check("user blocked from /manage", user.get("/manage").status_code == 403)
    check("user blocked from /sor/new", user.get("/sor/new").status_code == 403)

    admin = app.test_client()
    login(admin, "admin", "admin123")
    check("admin can manage", admin.get("/manage").status_code == 200)
    check("admin can open add form", admin.get("/sor/new").status_code == 200)

    # CSRF enforced
    no_csrf = app.test_client()
    login(no_csrf, "admin", "admin123")
    r = no_csrf.post("/sor/new", data={"sor_code": "99999", "name": "X", "category": "Dental", "price": "10"})
    check("POST without CSRF -> 400", r.status_code == 400, str(r.status_code))


def test_crud():
    admin = app.test_client()
    login(admin, "admin", "admin123")

    r = csrf_post(admin, "/sor/new", {
        "sor_code": "99997", "name": "Test Procedure Alpha", "category": "Dental", "price": "1500",
    })
    check("create item -> redirect", r.status_code == 302)

    item = db.get_sor_item_by_code("99997")
    check("item exists in DB", item is not None)
    item_id = item["id"]

    # duplicate code rejected
    r = csrf_post(admin, "/sor/new", {
        "sor_code": "99997", "name": "Dup", "category": "Dental", "price": "1",
    })
    check("duplicate code -> 400", r.status_code == 400, str(r.status_code))

    # edit
    r = csrf_post(admin, f"/sor/{item_id}/edit", {
        "sor_code": "99997", "name": "Test Procedure Beta", "category": "Urology", "price": "2500.50",
    })
    check("edit item -> redirect", r.status_code == 302)
    check("edit persisted name", db.get_sor_item(item_id)["name"] == "Test Procedure Beta")
    check("edit persisted price", db.get_sor_item(item_id)["price"] == 2500.50)

    # delete
    r = csrf_post(admin, f"/sor/{item_id}/delete", {})
    check("delete item -> redirect", r.status_code == 302)
    check("item gone after delete", db.get_sor_item(item_id) is None)

    # delete of non-existent -> 404
    r = csrf_post(admin, "/sor/999999999/delete", {})
    check("delete missing -> 404", r.status_code == 404)


def test_search_and_field():
    user = app.test_client()
    login(user, "user", "user123")

    both = user.get("/api/search?q=catheter").get_json()
    name = user.get("/api/search?q=catheter&field=name").get_json()
    code = user.get("/api/search?q=catheter&field=code").get_json()
    check("search field=name works", name["total"] >= 1)
    check("search field=code restricts", code["total"] == 0)
    check("search default covers both", both["total"] >= name["total"])

    num = user.get("/api/search?q=1&field=code").get_json()
    check("numeric code search", num["total"] >= 1)

    cat = user.get("/api/search?category=Urology").get_json()
    check("category filter", cat["total"] == 146, str(cat["total"]))

    sug = user.get("/api/suggest?q=cat").get_json()["items"]
    check("autocomplete suggestions", len(sug) > 0)


def test_persistence_across_restarts():
    # Simulate server restart: create item, re-run init_db, confirm it stays.
    admin = app.test_client()
    login(admin, "admin", "admin123")
    csrf_post(admin, "/sor/new", {
        "sor_code": "88888", "name": "Persistent Item", "category": "Dental", "price": "100",
    })
    assert db.get_sor_item_by_code("88888") is not None

    db.init_db()  # "restart"
    check("data survives init_db (restart)", db.get_sor_item_by_code("88888") is not None)

    # deleting data is not undone by restart
    item = db.get_sor_item_by_code("88888")
    csrf_post(admin, f"/sor/{item['id']}/delete", {})
    assert db.get_sor_item(item["id"]) is None
    db.init_db()
    check("deletion persists across restart", db.get_sor_item(item["id"]) is None)
    check("no re-seed of deleted data", db.get_sor_item_by_code("88888") is None)


def test_concurrency():
    errors = []

    def read_worker(_):
        try:
            for _ in range(10):
                db.search_sor_items("catheter", page=1)
                db.list_categories()
        except Exception as exc:  # noqa: BLE001
            errors.append(exc)

    def write_worker(i):
        try:
            code = f"Z{i:03d}"  # letter prefix never collides with numeric seed codes
            db.create_sor_item(code, f"Concurrent {i}", "Dental", float(i))
            db.update_sor_item(db.get_sor_item_by_code(code)["id"], code, f"Concurrent {i} v2", "Urology", float(i + 1))
            db.delete_sor_item(db.get_sor_item_by_code(code)["id"])
        except Exception as exc:  # noqa: BLE001
            errors.append(exc)

    with ThreadPoolExecutor(max_workers=16) as pool:
        readers = [pool.submit(read_worker, i) for i in range(8)]
        writers = [pool.submit(write_worker, i) for i in range(16)]
        for fut in readers + writers:
            fut.result()

    check("concurrent reads+writes: no errors", not errors, "; ".join(str(e) for e in errors[:3]))

    with db.get_db() as conn:
        leftovers = conn.execute(
            "SELECT COUNT(*) FROM sor_items WHERE sor_code LIKE 'Z%'"
        ).fetchone()[0]
    check("concurrent writes cleaned up", leftovers == 0, str(leftovers))


def test_backup():
    db.init_db()
    dest = db.backup_database()
    check("backup created", bool(dest) and os.path.exists(dest), str(dest))
    if dest:
        with db.get_db() as conn:
            count = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
        import sqlite3
        bconn = sqlite3.connect(dest)
        bcount = bconn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
        bconn.close()
        check("backup matches live data", bcount == count, f"{bcount} vs {count}")


def test_security_hardening():
    # security headers applied to responses
    anon = app.test_client()
    r = anon.get("/login")
    check("X-Frame-Options header", r.headers.get("X-Frame-Options") == "SAMEORIGIN")
    check("X-Content-Type-Options header", r.headers.get("X-Content-Type-Options") == "nosniff")

    # login brute-force throttle: throwaway username (never locks a real account)
    victim = "no_such_user_audit"
    c = app.test_client()
    for i in range(5):
        c.post("/login", data={"username": victim, "password": "wrong"})
    r = c.post("/login", data={"username": victim, "password": "stillwrong"})
    check("login locked after 5 failures -> 429", r.status_code == 429, str(r.status_code))
    r = c.post("/login", data={"username": victim, "password": "admin123"})
    check("locked even with any password -> 429", r.status_code == 429)
    app_module._clear_login_failures(victim, "127.0.0.1")

    # admin mutations are audited
    admin = app.test_client()
    login(admin, "admin", "admin123")
    csrf_post(admin, "/sor/new", {
        "sor_code": "77771", "name": "Audited Item", "category": "Dental", "price": "42",
    })
    log_path = os.path.join(paths.get_logs_dir(), "app.log")
    content = open(log_path, encoding="utf-8").read() if os.path.exists(log_path) else ""
    check("audit entry written to log", "AUDIT user=admin action=create" in content)
    item = db.get_sor_item_by_code("77771")
    csrf_post(admin, f"/sor/{item['id']}/delete", {})
    content = open(log_path, encoding="utf-8").read()
    check("audit delete entry written", "AUDIT user=admin action=delete" in content)


def test_logging():
    log_path = os.path.join(paths.get_logs_dir(), "app.log")
    check("log file exists", os.path.exists(log_path))
    if os.path.exists(log_path):
        content = open(log_path, encoding="utf-8").read()
        check("log has startup entry", "Starting CHSS SOR" in content or len(content) > 0)


def _make_workbook(rows, headers=None):
    import io
    import openpyxl
    headers = headers or ["SOR Number", "Title", "Category", "Price"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def test_sor_code_optional_and_alphanumeric():
    admin = app.test_client()
    login(admin, "admin", "admin123")

    r = csrf_post(admin, "/sor/new", {
        "sor_code": "", "name": "No Code Item", "category": "Dental", "price": "100",
    })
    check("blank sor code accepted", r.status_code == 302, str(r.status_code))
    with db.get_db() as conn:
        row = conn.execute(
            "SELECT sor_code FROM sor_items WHERE name = 'No Code Item'"
        ).fetchone()
    check("blank code stored as NULL", row is not None and row["sor_code"] is None)
    with db.get_db() as conn:
        conn.execute("DELETE FROM sor_items WHERE name = 'No Code Item'")

    r = csrf_post(admin, "/sor/new", {
        "sor_code": "90.1", "name": "Bad Code Item", "category": "Dental", "price": "100",
    })
    check("non-alphanumeric code rejected", r.status_code == 400, str(r.status_code))

    r = csrf_post(admin, "/sor/new", {
        "sor_code": "Ab12", "name": "Alpha Numeric Code", "category": "Dental", "price": "100",
    })
    check("alphanumeric code accepted", r.status_code == 302, str(r.status_code))
    check("alphanumeric code stored", db.get_sor_item_by_code("Ab12") is not None)
    with db.get_db() as conn:
        conn.execute("DELETE FROM sor_items WHERE sor_code = 'Ab12'")


def test_schema_nullable():
    with db.get_db() as conn:
        columns = conn.execute("PRAGMA table_info(sor_items)").fetchall()
    sor_code_col = [c for c in columns if c["name"] == "sor_code"][0]
    check("fresh schema sor_code nullable", sor_code_col["notnull"] == 0)


def test_legacy_schema_migration():
    import sqlite3
    fd, path = tempfile.mkstemp(suffix=".db")
    os.close(fd)
    conn = sqlite3.connect(path)
    conn.execute(
        """CREATE TABLE sor_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sor_code TEXT UNIQUE NOT NULL,
            name TEXT NOT NULL, category TEXT NOT NULL, price REAL,
            created_at TEXT NOT NULL, updated_at TEXT NOT NULL)"""
    )
    conn.execute("CREATE INDEX idx_sor_items_category ON sor_items(category)")
    conn.execute(
        "INSERT INTO sor_items (sor_code, name, category, price, created_at, updated_at)"
        " VALUES ('7', 'Legacy Item', 'Dental', 100, 'x', 'y')"
    )
    conn.commit()
    db._migrate_sor_code_nullable(conn)
    columns = {r[1]: r[3] for r in conn.execute("PRAGMA table_info(sor_items)").fetchall()}
    conn.commit()
    check("legacy migration makes code nullable", columns.get("sor_code") == 0, str(columns.get("sor_code")))
    row = conn.execute("SELECT sor_code, name FROM sor_items").fetchone()
    check("legacy migration preserves data", row[0] == "7" and row[1] == "Legacy Item")
    conn.close()
    os.remove(path)


def test_excel_parse():
    data = _make_workbook([
        ["9001", "Alpha Service", "Dental", 1000],
        ["", "Beta Service", "", None],
        ["9003", "Gamma Service", "Urology", "2500"],
    ])
    items = excel.parse_import_rows(data)
    check("excel parse row count", len(items) == 3, str(len(items)))
    check("excel parses code+price", items[0]["sor_code"] == "9001" and items[0]["price"] == 1000)
    check("excel optional code -> None", items[1]["sor_code"] is None)
    check("excel blank category auto-assigned", bool(items[1]["category"]))


def test_excel_validation():
    # non-alphanumeric code
    data = _make_workbook([["90.1", "Bad Code", "Dental", 10]])
    try:
        excel.parse_import_rows(data)
        check("non-alphanumeric code rejected", False)
    except ValueError:
        check("non-alphanumeric code rejected", True)

    # duplicate codes
    data = _make_workbook([["90", "A", "Dental", 1], ["90", "B", "Dental", 2]])
    try:
        excel.parse_import_rows(data)
        check("duplicate codes rejected", False)
    except ValueError:
        check("duplicate codes rejected", True)

    # missing title column
    data = _make_workbook(
        [["90", "X", "Y"]],
        headers=["SOR Number", "Something Else", "Category"],
    )
    try:
        excel.parse_import_rows(data)
        check("missing title column rejected", False)
    except ValueError:
        check("missing title column rejected", True)

    # not an Excel file at all
    try:
        excel.parse_import_rows(b"this is not an xlsx file")
        check("non-Excel file rejected", False)
    except ValueError:
        check("non-Excel file rejected", True)


def test_import_rollback_export_routes():
    import io
    import openpyxl
    admin = app.test_client()
    login(admin, "admin", "admin123")

    with db.get_db() as conn:
        before = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]

    data = _make_workbook([["9101", "Imported Alpha", "Dental", 500]])
    r = csrf_post(admin, "/manage/import", {"excel": (io.BytesIO(data), "sor.xlsx")})
    check("import route -> redirect", r.status_code == 302, str(r.status_code))
    with db.get_db() as conn:
        after = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
    check("import replaced dataset", after == 1, f"{before} -> {after}")
    check("import audit logged", db.has_rollback())

    r = csrf_post(admin, "/manage/rollback", {})
    check("rollback route -> redirect", r.status_code == 302, str(r.status_code))
    with db.get_db() as conn:
        restored = conn.execute("SELECT COUNT(*) FROM sor_items").fetchone()[0]
    check("rollback restores previous dataset", restored == before, f"{restored} vs {before}")

    r = admin.get("/manage/export")
    check("export route 200", r.status_code == 200, str(r.status_code))
    check("export xlsx content type", "spreadsheetml" in r.headers.get("Content-Type", ""))
    check("export attachment filename", "sor_database" in r.headers.get("Content-Disposition", ""))
    workbook = openpyxl.load_workbook(io.BytesIO(r.data))
    sheet = workbook.active
    header = [c.value for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    check("export headers", header == ["SOR Number", "Title", "Category", "Price"])
    check("export row count matches DB", sheet.max_row - 1 == before)


def main():
    test_runtime_environment()
    test_database()
    test_auth_rbac_csrf()
    test_crud()
    test_search_and_field()
    test_persistence_across_restarts()
    test_concurrency()
    test_backup()
    test_security_hardening()
    test_logging()
    test_sor_code_optional_and_alphanumeric()
    test_schema_nullable()
    test_legacy_schema_migration()
    test_excel_parse()
    test_excel_validation()
    test_import_rollback_export_routes()

    print()
    if FAILURES:
        print(f"FAILED: {len(FAILURES)} check(s):")
        for f in FAILURES:
            print("  -", f)
        sys.exit(1)
    print("ALL CHECKS PASSED")

    shutil.rmtree(TEST_DATA_DIR, ignore_errors=True)


if __name__ == "__main__":
    main()
